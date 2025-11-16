"""
apps/accounting/exports.py
Gestion des exportations comptables (CSV, PDF, Excel)
"""
import csv
from io import BytesIO
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Facture, Depense, CompteComptable, EcritureComptable,
    LigneEcriture
)
from .utils import ComptabiliteUtils


# ============================================================================
# UTILITAIRES
# ============================================================================
def get_csv_response(filename):
    """Crée une réponse HTTP pour CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff'.encode('utf8'))
    return response

def get_pdf_response(filename):
    """Crée une réponse HTTP pour PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def get_excel_response(filename):
    """Crée une réponse HTTP pour Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def create_pdf_styles():
    """Retourne les styles pour PDF"""
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name='CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=1
    ))

    styles.add(ParagraphStyle(
        name='CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=12
    ))

    return styles

def create_table_style():
    """Style de base pour tableaux PDF"""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2563eb')),
    ])


# ============================================================================
# EXPORT FACTURES
# ============================================================================
@login_required
def export_factures_csv(request):
    """Exporte les factures en CSV"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    factures = Facture.objects.filter(
        etablissement=request.user.etablissement
    ).select_related('apprenant', 'inscription')

    # Appliquer filtres
    statut = request.GET.get('statut')
    if statut:
        factures = factures.filter(statut=statut)

    response = get_csv_response(f'factures_{datetime.now().strftime("%Y%m%d")}.csv')
    writer = csv.writer(response)

    writer.writerow([
        'N° Facture', 'Date Émission', 'Date Échéance', 'Apprenant',
        'Type', 'Montant HT', 'TVA', 'Montant TTC', 'Montant Payé',
        'Solde', 'Statut'
    ])

    for facture in factures:
        writer.writerow([
            facture.numero_facture,
            facture.date_emission.strftime('%d/%m/%Y'),
            facture.date_echeance.strftime('%d/%m/%Y'),
            facture.apprenant.get_full_name(),
            facture.get_type_facture_display(),
            f"{facture.montant_ht:.2f}",
            f"{facture.montant_tva:.2f}",
            f"{facture.montant_ttc:.2f}",
            f"{facture.montant_paye:.2f}",
            f"{facture.solde_restant:.2f}",
            facture.get_statut_display()
        ])

    return response

@login_required
def export_factures_excel(request):
    """Exporte les factures en Excel"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    factures = Facture.objects.filter(
        etablissement=request.user.etablissement
    ).select_related('apprenant')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"

    # Style d'en-tête
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # En-têtes
    headers = [
        'N° Facture', 'Date Émission', 'Apprenant', 'Type',
        'Montant HT', 'TVA', 'TTC', 'Payé', 'Solde', 'Statut'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Données
    for row, facture in enumerate(factures, 2):
        ws.cell(row, 1, facture.numero_facture)
        ws.cell(row, 2, facture.date_emission.strftime('%d/%m/%Y'))
        ws.cell(row, 3, facture.apprenant.get_full_name())
        ws.cell(row, 4, facture.get_type_facture_display())
        ws.cell(row, 5, float(facture.montant_ht))
        ws.cell(row, 6, float(facture.montant_tva))
        ws.cell(row, 7, float(facture.montant_ttc))
        ws.cell(row, 8, float(facture.montant_paye))
        ws.cell(row, 9, float(facture.solde_restant))
        ws.cell(row, 10, facture.get_statut_display())

    # Ajuster largeurs
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = get_excel_response(f'factures_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.write(buffer.read())

    return response


# ============================================================================
# EXPORT DÉPENSES
# ============================================================================
@login_required
def export_depenses_csv(request):
    """Exporte les dépenses en CSV"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    depenses = Depense.objects.filter(
        etablissement=request.user.etablissement
    )

    response = get_csv_response(f'depenses_{datetime.now().strftime("%Y%m%d")}.csv')
    writer = csv.writer(response)

    writer.writerow([
        'N° Dépense', 'Date', 'Fournisseur', 'Catégorie',
        'Description', 'Montant', 'Mode Paiement', 'Statut'
    ])

    for depense in depenses:
        writer.writerow([
            depense.numero_depense,
            depense.date_depense.strftime('%d/%m/%Y'),
            depense.fournisseur,
            depense.get_categorie_display(),
            depense.description,
            f"{depense.montant:.2f}",
            depense.mode_paiement,
            depense.get_statut_display()
        ])

    return response


# ============================================================================
# EXPORT BALANCE
# ============================================================================
@login_required
def export_balance_excel(request):
    """Exporte la balance en Excel"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    if not date_debut or not date_fin:
        messages.error(request, "Dates requises")
        return redirect('dashboard:comptable_rapports')

    from datetime import datetime as dt
    date_debut = dt.strptime(date_debut, '%Y-%m-%d').date()
    date_fin = dt.strptime(date_fin, '%Y-%m-%d').date()

    balance = ComptabiliteUtils.generer_balance(
        request.user.etablissement,
        date_debut,
        date_fin
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"

    # En-tête
    ws['A1'] = f"BALANCE GÉNÉRALE - {request.user.etablissement.nom}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"

    # Colonnes
    headers = ['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde Débiteur', 'Solde Créditeur']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2563eb", fill_type="solid")

    # Données
    row = 5
    for ligne in balance:
        ws.cell(row, 1, ligne['numero_compte'])
        ws.cell(row, 2, ligne['libelle'])
        ws.cell(row, 3, float(ligne['debit']))
        ws.cell(row, 4, float(ligne['credit']))
        ws.cell(row, 5, float(ligne['solde_debiteur']))
        ws.cell(row, 6, float(ligne['solde_crediteur']))
        row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = get_excel_response(f'balance_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.write(buffer.read())

    return response

@login_required
def export_balance_pdf(request):
    """Exporte la balance en PDF"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    from datetime import datetime as dt
    date_debut = dt.strptime(date_debut, '%Y-%m-%d').date()
    date_fin = dt.strptime(date_fin, '%Y-%m-%d').date()

    balance = ComptabiliteUtils.generer_balance(
        request.user.etablissement,
        date_debut,
        date_fin
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = create_pdf_styles()

    # Titre
    elements.append(Paragraph(
        f"BALANCE GÉNÉRALE<br/>{request.user.etablissement.nom}",
        styles['CustomTitle']
    ))
    elements.append(Paragraph(
        f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
        styles['CustomHeading']
    ))
    elements.append(Spacer(1, 20))

    # Tableau
    data = [['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde Deb.', 'Solde Créd.']]

    for ligne in balance:
        data.append([
            ligne['numero_compte'],
            ligne['libelle'][:30],
            f"{ligne['debit']:,.0f}",
            f"{ligne['credit']:,.0f}",
            f"{ligne['solde_debiteur']:,.0f}",
            f"{ligne['solde_crediteur']:,.0f}"
        ])

    table = Table(data, colWidths=[3 * cm, 5 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm])
    table.setStyle(create_table_style())
    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = get_pdf_response(f'balance_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.write(pdf)

    return response


# ============================================================================
# EXPORT GRAND LIVRE
# ============================================================================
@login_required
def export_grand_livre_pdf(request):
    """Exporte le grand livre en PDF"""
    if request.user.role not in ['ADMIN', 'COMPTABLE']:
        messages.error(request, "Accès non autorisé")
        return redirect('dashboard:redirect')

    compte_id = request.GET.get('compte')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    if not all([compte_id, date_debut, date_fin]):
        messages.error(request, "Paramètres requis")
        return redirect('dashboard:comptable_rapports')

    from datetime import datetime as dt
    compte = CompteComptable.objects.get(id=compte_id)
    date_debut = dt.strptime(date_debut, '%Y-%m-%d').date()
    date_fin = dt.strptime(date_fin, '%Y-%m-%d').date()

    grand_livre = ComptabiliteUtils.generer_grand_livre(
        request.user.etablissement,
        compte,
        date_debut,
        date_fin
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = create_pdf_styles()

    # Titre
    elements.append(Paragraph(
        f"GRAND LIVRE - Compte {compte.numero_compte}",
        styles['CustomTitle']
    ))
    elements.append(Paragraph(
        f"{compte.libelle}",
        styles['CustomHeading']
    ))
    elements.append(Spacer(1, 20))

    # Solde initial
    elements.append(Paragraph(
        f"Solde initial: {grand_livre['solde_initial']:,.0f} FCFA",
        styles['CustomHeading']
    ))
    elements.append(Spacer(1, 10))

    # Mouvements
    data = [['Date', 'Pièce', 'Libellé', 'Débit', 'Crédit', 'Solde']]

    for mvt in grand_livre['mouvements'][:100]:
        data.append([
            mvt['date'].strftime('%d/%m/%Y'),
            mvt['numero_piece'],
            mvt['libelle'][:25],
            f"{mvt['debit']:,.0f}",
            f"{mvt['credit']:,.0f}",
            f"{mvt['solde']:,.0f}"
        ])

    table = Table(data, colWidths=[2 * cm, 2.5 * cm, 6 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    table.setStyle(create_table_style())
    elements.append(table)

    # Solde final
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Solde final: {grand_livre['solde_final']:,.0f} FCFA",
        styles['CustomHeading']
    ))

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = get_pdf_response(f'grand_livre_{compte.numero_compte}_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.write(pdf)

    return response
