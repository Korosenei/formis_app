from decimal import Decimal
from django.db.models import Sum, F, Q
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from io import BytesIO
import openpyxl
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ComptabiliteUtils:
    """Utilitaires pour les calculs comptables"""

    @staticmethod
    def calculer_solde_compte(compte, date_debut=None, date_fin=None):
        """Calcule le solde d'un compte pour une période donnée"""
        from .models import LigneEcriture

        queryset = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='VALIDEE'
        )

        if date_debut:
            queryset = queryset.filter(ecriture__date_ecriture__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(ecriture__date_ecriture__lte=date_fin)

        totaux = queryset.aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit')
        )

        debit = totaux['total_debit'] or Decimal('0.00')
        credit = totaux['total_credit'] or Decimal('0.00')

        # Le solde dépend du type de compte
        if compte.categorie in ['ACTIF', 'CHARGES']:
            return debit - credit
        else:  # PASSIF, PRODUITS, TRESORERIE
            return credit - debit

    @staticmethod
    def generer_balance(etablissement, date_debut, date_fin):
        """Génère la balance générale"""
        from .models import CompteComptable, LigneEcriture

        comptes = CompteComptable.objects.filter(
            etablissement=etablissement,
            est_actif=True
        ).order_by('numero_compte')

        balance = []

        for compte in comptes:
            lignes = LigneEcriture.objects.filter(
                compte=compte,
                ecriture__statut='VALIDEE',
                ecriture__date_ecriture__range=[date_debut, date_fin]
            )

            totaux = lignes.aggregate(
                total_debit=Sum('debit'),
                total_credit=Sum('credit')
            )

            debit = totaux['total_debit'] or Decimal('0.00')
            credit = totaux['total_credit'] or Decimal('0.00')

            if debit != 0 or credit != 0:
                solde = ComptabiliteUtils.calculer_solde_compte(compte, date_debut, date_fin)

                balance.append({
                    'numero_compte': compte.numero_compte,
                    'libelle': compte.libelle,
                    'debit': debit,
                    'credit': credit,
                    'solde_debiteur': solde if solde > 0 else Decimal('0.00'),
                    'solde_crediteur': abs(solde) if solde < 0 else Decimal('0.00'),
                })

        return balance

    @staticmethod
    def generer_grand_livre(etablissement, compte, date_debut, date_fin):
        """Génère le grand livre pour un compte"""
        from .models import LigneEcriture

        lignes = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='VALIDEE',
            ecriture__date_ecriture__range=[date_debut, date_fin]
        ).select_related('ecriture', 'ecriture__journal').order_by('ecriture__date_ecriture')

        # Solde initial
        solde_initial = ComptabiliteUtils.calculer_solde_compte(
            compte,
            date_fin=date_debut - timedelta(days=1)
        )

        mouvements = []
        solde_courant = solde_initial

        for ligne in lignes:
            if compte.categorie in ['ACTIF', 'CHARGES']:
                solde_courant += ligne.debit - ligne.credit
            else:
                solde_courant += ligne.credit - ligne.debit

            mouvements.append({
                'date': ligne.ecriture.date_ecriture,
                'numero_piece': ligne.ecriture.numero_piece,
                'libelle': ligne.libelle,
                'journal': ligne.ecriture.journal.libelle,
                'debit': ligne.debit,
                'credit': ligne.credit,
                'solde': solde_courant,
            })

        return {
            'compte': compte,
            'solde_initial': solde_initial,
            'mouvements': mouvements,
            'solde_final': solde_courant,
        }


class RapportComptablePDF:
    """Générateur de rapports comptables en PDF"""

    @staticmethod
    def generer_facture(facture):
        """Génère une facture en PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # En-tête
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30,
            alignment=1
        )

        elements.append(Paragraph(f"FACTURE N° {facture.numero_facture}", title_style))
        elements.append(Spacer(1, 20))

        # Informations établissement et client
        info_data = [
            ['Établissement:', facture.etablissement.nom],
            ['Adresse:', facture.etablissement.adresse],
            ['Téléphone:', facture.etablissement.telephone or ''],
            ['', ''],
            ['Client:', facture.apprenant.get_full_name()],
            ['Matricule:', facture.apprenant.matricule],
            ['Date d\'émission:', facture.date_emission.strftime('%d/%m/%Y')],
            ['Date d\'échéance:', facture.date_echeance.strftime('%d/%m/%Y')],
        ]

        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 30))

        # Lignes de la facture
        lignes_data = [['Description', 'Quantité', 'Prix unitaire', 'Montant']]

        for ligne in facture.lignes.all():
            lignes_data.append([
                ligne.description,
                f"{ligne.quantite}",
                f"{ligne.prix_unitaire:,.0f} FCFA",
                f"{ligne.montant:,.0f} FCFA"
            ])

        lignes_table = Table(lignes_data, colWidths=[3 * inch, 1 * inch, 1.5 * inch, 1.5 * inch])
        lignes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(lignes_table)
        elements.append(Spacer(1, 20))

        # Totaux
        totaux_data = [
            ['Montant HT:', f"{facture.montant_ht:,.0f} FCFA"],
            ['TVA ({:.1f}%):'.format(facture.taux_tva), f"{facture.montant_tva:,.0f} FCFA"],
            ['TOTAL TTC:', f"{facture.montant_ttc:,.0f} FCFA"],
        ]

        totaux_table = Table(totaux_data, colWidths=[4.5 * inch, 2 * inch])
        totaux_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -2), 'Helvetica', 11),
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 14),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#2563eb')),
        ]))

        elements.append(totaux_table)

        # Notes
        if facture.description:
            elements.append(Spacer(1, 30))
            elements.append(Paragraph(f"<b>Notes:</b> {facture.description}", styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generer_balance(balance_data, etablissement, date_debut, date_fin):
        """Génère la balance générale en PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # En-tête
        title = f"BALANCE GÉNÉRALE<br/>{etablissement.nom}"
        periode = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"

        elements.append(Paragraph(title, styles['Title']))
        elements.append(Paragraph(periode, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Données de la balance
        data = [['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde déb.', 'Solde créd.']]

        total_debit = Decimal('0.00')
        total_credit = Decimal('0.00')
        total_solde_deb = Decimal('0.00')
        total_solde_cred = Decimal('0.00')

        for ligne in balance_data:
            data.append([
                ligne['numero_compte'],
                ligne['libelle'][:30],
                f"{ligne['debit']:,.0f}",
                f"{ligne['credit']:,.0f}",
                f"{ligne['solde_debiteur']:,.0f}",
                f"{ligne['solde_crediteur']:,.0f}",
            ])
            total_debit += ligne['debit']
            total_credit += ligne['credit']
            total_solde_deb += ligne['solde_debiteur']
            total_solde_cred += ligne['solde_crediteur']

        # Ligne de totaux
        data.append([
            'TOTAL',
            '',
            f"{total_debit:,.0f}",
            f"{total_credit:,.0f}",
            f"{total_solde_deb:,.0f}",
            f"{total_solde_cred:,.0f}",
        ])

        table = Table(data, colWidths=[0.8 * inch, 2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
            ('FONT', (0, 1), (-1, -2), 'Helvetica', 8),
            ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
        ]))

        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer


class ExportComptableExcel:
    """Export de données comptables en Excel"""

    @staticmethod
    def exporter_balance(balance_data, etablissement, date_debut, date_fin):
        """Exporte la balance en Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance générale"

        # En-tête
        ws['A1'] = f"BALANCE GÉNÉRALE - {etablissement.nom}"
        ws['A2'] = f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"

        # Mise en forme de l'en-tête
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'].font = Font(size=11)

        # En-têtes de colonnes
        headers = ['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde débiteur', 'Solde créditeur']
        ws.append([])  # Ligne vide
        ws.append(headers)

        # Mise en forme des en-têtes
        header_row = ws[4]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données
        for ligne in balance_data:
            ws.append([
                ligne['numero_compte'],
                ligne['libelle'],
                float(ligne['debit']),
                float(ligne['credit']),
                float(ligne['solde_debiteur']),
                float(ligne['solde_crediteur']),
            ])

        # Format des nombres
        for row in ws.iter_rows(min_row=5, max_col=6):
            for cell in row[2:]:
                cell.number_format = '#,##0.00'

        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer